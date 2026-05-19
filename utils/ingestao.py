import requests
from bs4 import BeautifulSoup

from utils.documentos import chunk_texto, salvar_chunks


def processar_url(url: str) -> dict:
    try:
        resp = requests.get(
            url, timeout=30,
            headers={"User-Agent": "Mozilla/5.0"},
        )
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")
        titulo = soup.title.string.strip() if soup.title else url
        for tag in soup(["script", "style", "nav", "footer", "header", "aside"]):
            tag.decompose()
        texto = soup.get_text(separator="\n", strip=True)
        if not texto:
            return {"status": "erro", "mensagem": "Nenhum texto extraído da URL."}
        chunks = chunk_texto(texto)
        total = salvar_chunks(
            chunks,
            documento_id=f"url_{url}",
            extra_metadata={"fonte": "url", "url": url, "titulo": titulo},
        )
        return {
            "status": "ok",
            "total_chunks": total,
            "total_caracteres": len(texto),
            "titulo": titulo,
            "url": url,
        }
    except requests.RequestException as e:
        return {"status": "erro", "mensagem": f"Erro ao acessar URL: {e}"}
    except Exception as e:
        return {"status": "erro", "mensagem": str(e)}


def processar_html(html_bytes: bytes, nome_arquivo: str = "") -> dict:
    try:
        soup = BeautifulSoup(html_bytes, "html.parser")
        titulo = soup.title.string.strip() if soup.title else nome_arquivo
        for tag in soup(["script", "style", "nav", "footer", "header", "aside"]):
            tag.decompose()
        texto = soup.get_text(separator="\n", strip=True)
        if not texto:
            return {"status": "erro", "mensagem": "Nenhum texto extraído do HTML."}
        chunks = chunk_texto(texto)
        total = salvar_chunks(
            chunks,
            documento_id=nome_arquivo or "html",
            extra_metadata={"fonte": "html", "arquivo": nome_arquivo, "titulo": titulo},
        )
        return {
            "status": "ok",
            "total_chunks": total,
            "total_caracteres": len(texto),
            "titulo": titulo,
            "arquivo": nome_arquivo,
        }
    except Exception as e:
        return {"status": "erro", "mensagem": str(e)}


def processar_instagram(perfil: str, max_posts: int = 10) -> dict:
    try:
        import instaloader
        L = instaloader.Instaloader(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            max_connection_attempts=1,
        )
        try:
            import streamlit as st
            ig_user = st.secrets.get("INSTAGRAM_USER")
            ig_pass = st.secrets.get("INSTAGRAM_PASS")
            if ig_user and ig_pass:
                L.login(ig_user, ig_pass)
        except Exception:
            pass
        profile = instaloader.Profile.from_username(L.context, perfil)

        partes = [
            f"Perfil Instagram: {profile.full_name} (@{profile.username})",
            f"Seguidores: {profile.followers}",
            f"Seguindo: {profile.followees}",
            f"Publicações: {profile.mediacount}",
        ]
        if profile.biography:
            partes.append(f"Biografia: {profile.biography}")

        posts_text = []
        for i, post in enumerate(profile.get_posts()):
            if i >= max_posts:
                break
            legenda = post.caption or ""
            hashtags = " ".join(f"#{t}" for t in post.caption_hashtags) if post.caption_hashtags else ""
            partes_post = [
                f"\n--- Post {i+1} ---",
                f"Data: {post.date}",
                f"Curtidas: {post.likes}",
                f"Comentários: {post.comments}",
            ]
            if legenda:
                partes_post.append(f"Legenda: {legenda[:2000]}")
            if hashtags:
                partes_post.append(f"Hashtags: {hashtags}")
            posts_text.append("\n".join(partes_post))

        if posts_text:
            partes.append("Posts recentes:\n" + "\n\n".join(posts_text))

        texto = "\n\n".join(partes)
        if not texto.strip():
            return {"status": "erro", "mensagem": "Nenhum conteúdo extraído do perfil."}

        chunks = chunk_texto(texto)
        total = salvar_chunks(
            chunks,
            documento_id=f"ig_{perfil}",
            extra_metadata={"fonte": "instagram", "perfil": perfil},
        )
        return {
            "status": "ok",
            "total_chunks": total,
            "total_caracteres": len(texto),
            "perfil": perfil,
            "posts": min(max_posts, i + 1),
        }
    except instaloader.exceptions.ProfileNotExistsException:
        return {"status": "erro", "mensagem": f"Perfil @{perfil} não encontrado ou Instagram bloqueou a requisição. Tente adicionar INSTAGRAM_USER e INSTAGRAM_PASS nos Secrets do Streamlit Cloud."}
    except instaloader.exceptions.ConnectionException as e:
        return {"status": "erro", "mensagem": f"Erro de conexão com Instagram: {e}. Se estiver no Streamlit Cloud, o IP pode estar bloqueado — tente adicionar INSTAGRAM_USER e INSTAGRAM_PASS nos Secrets."}
    except ImportError:
        return {"status": "erro", "mensagem": "instaloader não instalado. Execute: pip install instaloader"}
    except Exception as e:
        return {"status": "erro", "mensagem": str(e)}


def processar_planilha(arquivo_bytes: bytes, nome_arquivo: str = "planilha.xlsx") -> dict:
    try:
        import openpyxl
        import io
        wb = openpyxl.load_workbook(io.BytesIO(arquivo_bytes), data_only=True)
        partes = [f"Planilha: {nome_arquivo}", f"Abas: {wb.sheetnames}"]
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            linhas_texto = []
            for row in ws.iter_rows(values_only=True):
                celulas = [str(c).strip() for c in row if c is not None]
                if celulas:
                    linhas_texto.append(" | ".join(celulas))
            if linhas_texto:
                partes.append(f"\n--- {sheet_name} ---\n" + "\n".join(linhas_texto))
        texto = "\n\n".join(partes)
        if not texto.strip():
            return {"status": "erro", "mensagem": "Nenhum conteúdo extraído da planilha."}
        import time
        documento_id = f"xls_{int(time.time())}"
        chunks = chunk_texto(texto)
        total = salvar_chunks(
            chunks,
            documento_id=documento_id,
            extra_metadata={"fonte": "planilha", "arquivo": nome_arquivo, "titulo": nome_arquivo},
        )
        return {
            "status": "ok",
            "total_chunks": total,
            "total_caracteres": len(texto),
            "titulo": nome_arquivo,
        }
    except ImportError:
        return {"status": "erro", "mensagem": "openpyxl não instalado. Execute: pip install openpyxl"}
    except Exception as e:
        return {"status": "erro", "mensagem": str(e)}


def processar_texto(texto: str, titulo: str = "") -> dict:
    try:
        if not texto.strip():
            return {"status": "erro", "mensagem": "Nenhum texto informado."}
        titulo = titulo.strip() or "Texto colado"
        import time
        documento_id = f"texto_{int(time.time())}"
        chunks = chunk_texto(texto)
        total = salvar_chunks(
            chunks,
            documento_id=documento_id,
            extra_metadata={"fonte": "texto", "arquivo": titulo, "titulo": titulo},
        )
        return {
            "status": "ok",
            "total_chunks": total,
            "total_caracteres": len(texto),
            "titulo": titulo,
        }
    except Exception as e:
        return {"status": "erro", "mensagem": str(e)}
